# choice = input("1 for guess game 2 for car game ")
#
# if choice == "1":
#     secret_number = 9
#     guess_count = 0
#     guess_limit = 3
#     while guess_count < guess_limit:
#         guess = int(input("Guess the number: "))
#         guess_count+=1
#         if guess == secret_number:
#             print("Done")
#             break
#     else:
#         print("no more attempts")
#
# if choice == "2":
#     command = ""
#     started = False
#     while True:
#         command = input(">").lower()
#         if command == "start":
#             if started:
#                 print("already started")
#             else:
#                 started=True
#             print("car started")
#         elif command == "stop":
#             if not started:
#                 print("already stopped")
#             else:
#                 started = False
#             print("car stopped")
#         elif command == "help":
#             print("""
# start
# stop
# quit
#             """)
#         elif command == "quit":
#             break
#         else:
#             print("invalid command")

# FOR LOOP
# prices = [1,2,3]
# sum = 0
# for price in prices:
#     sum += price
# print(f"Total: {sum}")

# NESTED LOOP

# for x in range(4):  # excludes this digit
#     for y in range(2):
#         print(f"{x}, {y}")
#
# PRINT AN F
# x_per_row = [5, 2, 5, 2, 2]
# for x_count in x_per_row:
#     output = ""
#     for count in range(x_count):
#         output += "x"
#     print(output)

# FIND BIGGEST N
# numbers = [1, 3, 3, 3, 2, 4, 5]
# max = numbers[0]
# for n in numbers:
#     if n > max:
#         max = n
# print(max)

# 2D LISTS
# matrix = [
#     [1,2,3],
#     [4,5,6],
#     [7,8,9]
# ]
# print(matrix[0][0])
#
# for row in matrix:
#     for item in row:
#         print(item)
#
# numbers.append(13)
# numbers.insert(0, 0)
# numbers.remove(3)
# numbers.pop()  # removes last item
# print(50 in numbers)
# print(numbers.count(1))
# numbers.sort()
# numbers.reverse()
# numbers2 = numbers.copy()
# print(numbers2)
#
# uniques = []
# for number in numbers:
#     if number not in uniques:
#         uniques.append(number)
# print(uniques)
#
# unpacking
# coor = (1, 2, 3)
# x, y, z = coor
#
# dictionaries
# customer = {  # keys should be unique
#     "name": "Doe",
#     "age": 30,
#     "bool": True
# }
# print(customer["name"])
# print(customer.get("default", "default"))
#
# phone = input("phone")
# digits_mapping = {
#     "1": "one",
#     "2": "two",
#     "3": "three",
#     "4": "four"
# }
# output = ""
# for ch in phone:
#     output += digits_mapping.get(ch, "?")+" "
# print(output)

# EMOJI
def emoji_conv(msg):
    split = msg.split(" ")
    emojis_dict = {
        ":)": "smile",
        ":(": "sad"
    }
    output = ""
    for word in split:
        output += emojis_dict.get(word, word)
    return output


# msgg = input(">")
# print(emoji_conv(msgg))

# def greet_user(name, last):
#     print("Hi")
#     print(f"Hi {name} {last}")
# greet_user("john", last="doe")

# def square(n):
#     return n*n
# print(square(2))

# try:
#     age = int(input("age "))
#     income = 200
#     risk = income / age
#     print(age)
# except ValueError:
#     print("invalid")
# except ZeroDivisionError:
#     print("age cannot be 0")

# class Point:
#     def __init__(self,x ,y):
#         self.x = x
#         self.y = y
#     def move(self):
#         print("move")
#
#     def draw(self):
#         print("draw")


# class Person:
#     def __init__(self, name):
#         self.name = name
#
#     def talk(self):
#         print(f"Hi my name is {self.name}")


# person1 = Person("Jonn")
# person1.talk()


# class Animal:
#     def walk(self):
#         print("walk")


import utils
from converters import kg_to_lbs


# class Dog(Animal):  # inheritance, use pass for an empty class
#     def bark(self):
#         print("bark")


# print(kg_to_lbs(12))
# print(utils.find_max([1,2,3,4]))

# import ecommerce.shipping
from ecommerce import shipping
import random

# team = ["a","b","c","d"]
# print(random.choice(team))

# for i in range(3):
#     print(random.randint(10,20))


# class Dice():
#     def roll(self):
#         first = random.randint(1, 6)
#         second = random.randint(1, 6)
#         print(first, second)


# dice = Dice()
# dice.roll()

from pathlib import Path

# path = Path()
# for file in path.glob("*.py"):
#     print(file)

##################################################

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Load the workbook
wb = xl.load_workbook('transactionspy.xlsx')

# List all sheet names
print(wb.sheetnames)

# Assuming you want to access the first sheet
sheet = wb[wb.sheetnames[0]]  # HOJA 1

# Access a specific cell
# cell = sheet.cell(1, 1)


def process_workbook(file_name):
    wb = xl.load_workbook(file_name)
    sheet = wb[wb.sheetnames[0]]

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(file_name)
